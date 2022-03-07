
from importlib.resources import path
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium import webdriver
from bs4 import BeautifulSoup
import time
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
import os

class Excel_handler:
    template_path : str = 'prijsbepaling_template.xlsx'
    sheet_name : str= 'Blad1'
    username :str = os.environ.get('USERNAME')
    
    def __init__(self):
        self.wb = load_workbook(self.template_path)
        self.working_sheet = self.wb[self.sheet_name]



    def concat_list_to_dataframe(self,*arrays:list)->pd.DataFrame:
        """concatane the arrays in 1 pandas dataframe"""
        print('adding to excel')
        df = pd.DataFrame()
        for array in arrays:
            temp_df = pd.DataFrame(array)
            df = pd.concat([temp_df,df])
        print('added')
        return df
        


    def add_dataframe_to_excel_and_save(self,dataFrame:pd.DataFrame):
        """add the scrape data to the template provided"""
        begin_cell = 8
        omschrijving_cell = 1
        price_cell = 2
        amount_cell = 3
        for i in range(len(dataFrame)):
            self.working_sheet.cell(i+begin_cell,omschrijving_cell).value = dataFrame.iloc[i,0]
            self.working_sheet.cell(i+begin_cell,price_cell).value = dataFrame.iloc[i,1]
            self.working_sheet.cell(i+begin_cell,amount_cell).value = 0
        
        print('saving...')
        path = f'C:\\Users\\{self.username}\\Desktop\\excel_template.xlsx'
        self.wb.save(path)
        print('saved.')



class Data:
    reloadTime : int
    
    colmuns : list = ['name','exclusief-BTW']
    def __init__(self,reloadTime:int):
        """Needs to recieve the reload time=> faster internet and PC equals less reload time
            Needs to recieve the amount of pages per material , for wood it is 14"""
        self.reloadTime = reloadTime
        

    def get_wood(self,driver:webdriver,total_pages:int,material:str)->list:
        '''Loops through all the pages to get all the wood articles'''
        print(f'getting {material}...')
        wood : list = []
        for page in tqdm(range(1,total_pages+1)):
            URL = {
                'wood':f'https://www.martenshout.be/nl/collection/hout?page={page}&numberOfItems=12&sortBy=title&sorting=ASC',
                'platen':f'https://www.martenshout.be/nl/collection/constructieplaten?page={page}&numberOfItems=12&sortBy=title&sorting=ASC',
                'fineer':f'https://www.martenshout.be/nl/collection/fineer?page={page}&numberOfItems=12&sortBy=title&sorting=ASC'
            }
            driver.get(URL.get(material))
            #sleep so the page can load, hardcoded: this can be a problem later on
            time.sleep(self.reloadTime)
            soup = BeautifulSoup(driver.page_source,'html.parser')
            
            for div in soup.find_all('div',class_='product-list-item col-lg-4 col-md-6 col-sm-12 mb-5'):
                temp_list = []
                name_product =div.find('span').text
                exclusief = self.split_currencies(div.find('div',class_='price').text)
                temp_list.append(name_product)
                temp_list.append(exclusief)
                wood.append(temp_list)
        driver.close()
        
        return wood
 

    def split_currencies(self,raw_text:str):
        """Splits raw prices into: inclusief btw per eenheid AND exclusief btw"""
        arr = []
        exclusief_index = 1
        #getting rid of the noice
        raw_text = raw_text.replace(u'\xa0', u' ')
        arr = raw_text.split(' ')
        currencie = arr[exclusief_index].replace('.','')
        currencie = float(currencie.replace(',','.'))

        return currencie




def main():
    #initialize the chromedriver
    options = Options()
    options.headless = True
    
    
    #make object from Data
    data= Data(reloadTime=3)
    excel_handler = Excel_handler()
    #get wood 
    
    wood = data.get_wood(webdriver.Firefox(options=options),total_pages=14,material='wood')
    
    #get platen
    construtie_platen = data.get_wood(webdriver.Firefox(options=options),total_pages=12,material='platen')
    
    #get fineer
    fineer = data.get_wood(webdriver.Firefox(options=options),total_pages=2,material='fineer')
    
    
    df = excel_handler.concat_list_to_dataframe(fineer,construtie_platen,wood)
    excel_handler.add_dataframe_to_excel_and_save(df)


if __name__ == '__main__':
    main()
